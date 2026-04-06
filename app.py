import streamlit as st
import pandas as pd
import io
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Transpositor Excel",
    page_icon="⇄",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Estilos ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}

/* Fondo general */
.stApp {
    background-color: #0f0f13;
    color: #e8e6f0;
}

/* Sidebar */
section[data-testid="stSidebar"] {
    background-color: #16151e;
    border-right: 1px solid #2a2838;
}
section[data-testid="stSidebar"] .stMarkdown p,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] .stRadio label {
    color: #b8b4cc !important;
    font-size: 0.82rem;
}

/* Título sidebar */
.sidebar-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.7rem;
    font-weight: 500;
    letter-spacing: 0.15em;
    color: #6c5ecf !important;
    text-transform: uppercase;
    margin-bottom: 0.5rem;
    padding-bottom: 0.5rem;
    border-bottom: 1px solid #2a2838;
}

/* Encabezado principal */
.main-header {
    display: flex;
    align-items: baseline;
    gap: 1rem;
    margin-bottom: 0.25rem;
}
.main-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.6rem;
    font-weight: 500;
    color: #e8e6f0;
    letter-spacing: -0.02em;
}
.main-tag {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.7rem;
    color: #6c5ecf;
    letter-spacing: 0.1em;
    background: #1e1b2e;
    padding: 2px 8px;
    border-radius: 2px;
    border: 1px solid #3a3560;
}
.main-sub {
    font-size: 0.88rem;
    color: #6b6880;
    margin-bottom: 2rem;
}

/* Cards / paneles */
.panel-card {
    background: #16151e;
    border: 1px solid #2a2838;
    border-radius: 6px;
    padding: 1.25rem 1.5rem;
    margin-bottom: 1rem;
}
.panel-label {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.65rem;
    letter-spacing: 0.12em;
    color: #6c5ecf;
    text-transform: uppercase;
    margin-bottom: 0.75rem;
}

/* Inputs */
.stTextInput input, .stNumberInput input, .stSelectbox select {
    background-color: #1e1b2e !important;
    border: 1px solid #2a2838 !important;
    border-radius: 4px !important;
    color: #e8e6f0 !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.85rem !important;
}
.stTextInput input:focus, .stNumberInput input:focus {
    border-color: #6c5ecf !important;
    box-shadow: 0 0 0 1px #6c5ecf33 !important;
}

/* Multiselect */
.stMultiSelect [data-baseweb="tag"] {
    background-color: #2a2060 !important;
    border: 1px solid #6c5ecf !important;
    border-radius: 2px !important;
}
.stMultiSelect [data-baseweb="tag"] span {
    color: #c4bfee !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.78rem !important;
}

/* Botones */
.stButton > button {
    background: #6c5ecf !important;
    color: #fff !important;
    border: none !important;
    border-radius: 4px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.82rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.05em !important;
    padding: 0.5rem 1.5rem !important;
    transition: background 0.15s !important;
}
.stButton > button:hover {
    background: #5a4db8 !important;
}

/* Download button */
.stDownloadButton > button {
    background: #1e3a2f !important;
    color: #4ade80 !important;
    border: 1px solid #2d5a42 !important;
    border-radius: 4px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.82rem !important;
    letter-spacing: 0.05em !important;
}
.stDownloadButton > button:hover {
    background: #2a4d3c !important;
}

/* Radio */
.stRadio [data-baseweb="radio"] span {
    color: #b8b4cc !important;
    font-size: 0.85rem !important;
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    background: transparent !important;
    border-bottom: 1px solid #2a2838 !important;
    gap: 0 !important;
}
.stTabs [data-baseweb="tab"] {
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.78rem !important;
    letter-spacing: 0.08em !important;
    color: #6b6880 !important;
    background: transparent !important;
    border-bottom: 2px solid transparent !important;
    padding: 0.5rem 1.25rem !important;
}
.stTabs [aria-selected="true"] {
    color: #c4bfee !important;
    border-bottom-color: #6c5ecf !important;
}

/* DataFrame */
.stDataFrame {
    border: 1px solid #2a2838 !important;
    border-radius: 4px !important;
}

/* Alerts */
.stSuccess {
    background: #0d2b1f !important;
    border: 1px solid #1a4a32 !important;
    color: #4ade80 !important;
    border-radius: 4px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.82rem !important;
}
.stError {
    background: #2b0d0d !important;
    border: 1px solid #4a1a1a !important;
    border-radius: 4px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.82rem !important;
}
.stInfo {
    background: #0d1a2b !important;
    border: 1px solid #1a324a !important;
    border-radius: 4px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.82rem !important;
}
.stWarning {
    background: #2b1f0d !important;
    border: 1px solid #4a3a1a !important;
    border-radius: 4px !important;
}

/* Expander */
.streamlit-expanderHeader {
    background: #16151e !important;
    border: 1px solid #2a2838 !important;
    border-radius: 4px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.78rem !important;
    color: #b8b4cc !important;
}

/* Métricas */
[data-testid="metric-container"] {
    background: #16151e;
    border: 1px solid #2a2838;
    border-radius: 6px;
    padding: 0.75rem 1rem;
}
[data-testid="metric-container"] label {
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.68rem !important;
    letter-spacing: 0.1em !important;
    color: #6c5ecf !important;
    text-transform: uppercase !important;
}
[data-testid="stMetricValue"] {
    font-family: 'IBM Plex Mono', monospace !important;
    color: #e8e6f0 !important;
}
</style>
""", unsafe_allow_html=True)


# ── Funciones ─────────────────────────────────────────────────────────────────

def get_sheet_names(file_obj):
    """Obtiene lista de hojas visibles del archivo Excel"""
    try:
        xl = pd.ExcelFile(file_obj)
        return xl.sheet_names
    except Exception as e:
        st.error(f"Error al leer hojas del archivo: {e}")
        return []


def get_preview_info(file_obj, sheet_name):
    """Obtiene información de previsualización de una hoja sin cargarla completa"""
    try:
        df_sample = pd.read_excel(file_obj, sheet_name=sheet_name, nrows=5)
        total_rows = len(pd.read_excel(file_obj, sheet_name=sheet_name))
        return {
            "rows": total_rows,
            "cols": len(df_sample.columns),
            "columns": list(df_sample.columns)[:5]  # primeras 5 columnas
        }
    except:
        return None


def process_dataframe(df, mode, params):
    """Procesa el DataFrame según el modo seleccionado"""
    if mode == "Ancho → Largo (melt)":
        id_c = params.get("id_cols", [])
        val_c = params.get("val_cols")
        if not val_c:
            raise ValueError("Debes seleccionar al menos una columna a trasponer")
        return df.melt(
            id_vars=id_c if id_c else None,
            value_vars=val_c,
            var_name=params.get("var_name", "Variable"),
            value_name=params.get("val_name", "Valor")
        )

    elif mode == "Largo → Ancho (pivot)":
        pivot_idx = params.get("pivot_index")
        pivot_col = params.get("pivot_cols")
        pivot_val = params.get("pivot_vals")
        agg_fn = params.get("agg_func", "sum")
        
        if not all([pivot_idx, pivot_col, pivot_val]):
            raise ValueError("Debes especificar índice, columnas y valores para el pivot")
        
        result = df.pivot_table(
            index=pivot_idx,
            columns=pivot_col,
            values=pivot_val,
            aggfunc=agg_fn
        ).reset_index()
        
        fill_v = params.get("fill_val", "0")
        if fill_v:
            try:
                fill_v = float(fill_v) if "." in fill_v else int(fill_v)
            except:
                pass
            result = result.fillna(fill_v)
        
        result.columns.name = None
        return result

    elif mode == "Transponer todo (.T)":
        val_c = params.get("val_cols")
        df_subset = df[val_c] if val_c else df
        return df_subset.T

    else:
        raise ValueError(f"Modo '{mode}' no implementado")


def apply_excel_styling(wb, sheet_name, do_headers=True, do_autofit=True, do_freeze=False):
    """Aplica estilos a la hoja Excel"""
    ws = wb[sheet_name]
    
    if do_headers and ws.max_row > 0:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6c5ecf", end_color="6c5ecf", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    if do_autofit:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            adjusted = min(max_len + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted
    
    if do_freeze and ws.max_row > 1:
        ws.freeze_panes = "A2"


def to_excel_bytes(df, sheet_name="Sheet1", index=False, 
                   do_headers=True, do_autofit=True, do_freeze=False):
    """Convierte DataFrame a bytes de Excel con estilos"""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index)
    buf.seek(0)
    wb = load_workbook(buf)
    apply_excel_styling(wb, sheet_name, do_headers, do_autofit, do_freeze)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def detect_numeric_columns(df):
    """Detecta columnas numéricas en el DataFrame"""
    return df.select_dtypes(include=['number']).columns.tolist()


def apply_advanced_filters(df, filters):
    """Aplica múltiples filtros al DataFrame"""
    df_filtered = df.copy()
    for filt in filters:
        if filt['col'] and filt['val']:
            try:
                if filt['op'] in ['contains', 'not_contains']:
                    if filt['op'] == 'contains':
                        df_filtered = df_filtered[df_filtered[filt['col']].astype(str).str.contains(filt['val'], na=False)]
                    else:
                        df_filtered = df_filtered[~df_filtered[filt['col']].astype(str).str.contains(filt['val'], na=False)]
                else:
                    val = float(filt['val']) if filt['val'].replace(".", "").replace("-", "").isdigit() else filt['val']
                    df_filtered = df_filtered[df_filtered[filt['col']].apply(
                        lambda x: eval(f"x {filt['op']} val", {"x": x, "val": val})
                    )]
            except Exception as e:
                st.warning(f"Filtro no aplicado en columna {filt['col']}: {e}")
    return df_filtered


# ── UI ────────────────────────────────────────────────────────────────────────

st.markdown("""
<div class="main-header">
  <span class="main-title">⇄ Transpositor Excel</span>
  <span class="main-tag">v2.0</span>
</div>
<p class="main-sub">Transpone, pivotea y reestructura hojas Excel sin código — descarga el resultado al instante.</p>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="sidebar-title">📁 Archivo de entrada</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Sube tu archivo Excel", type=["xlsx", "xls"],
                                label_visibility="collapsed")

    if uploaded is not None:
        # Detectar hojas disponibles
        sheet_names = get_sheet_names(uploaded)
        
        if sheet_names:
            st.markdown("---")
            st.markdown('<div class="sidebar-title">📋 Selección de hoja</div>', unsafe_allow_html=True)
            
            # Mostrar información de hojas disponibles
            with st.expander(f"📊 {len(sheet_names)} hoja(s) disponible(s)", expanded=False):
                for idx, sname in enumerate(sheet_names, 1):
                    info = get_preview_info(uploaded, sname)
                    if info:
                        st.caption(f"**{idx}. {sname}**  \n{info['rows']:,} filas × {info['cols']} cols")
                    else:
                        st.caption(f"**{idx}. {sname}**")
            
            sheet_input = st.selectbox(
                "Hoja a procesar",
                sheet_names,
                label_visibility="collapsed"
            )
        else:
            sheet_input = st.text_input("Nombre de hoja", value="Sheet1")

        st.markdown("---")
        st.markdown('<div class="sidebar-title">⚙️ Parámetros de lectura</div>', unsafe_allow_html=True)
        header_row  = st.number_input("Fila de encabezados", min_value=1, value=1, step=1)
        skip_rows   = st.number_input("Saltar filas al inicio", min_value=0, value=0, step=1)
        nrows_val   = st.number_input("Leer máx. N filas (0 = todo)", min_value=0, value=0, step=1)

        st.markdown("---")
        st.markdown('<div class="sidebar-title">💾 Archivo de salida</div>', unsafe_allow_html=True)
        out_filename = st.text_input("Nombre archivo resultado", value="resultado.xlsx")
        out_sheet    = st.text_input("Nombre hoja resultado", value="Resultado")
        include_idx  = st.checkbox("Incluir índice", value=False)

        st.markdown("---")
        st.markdown('<div class="sidebar-title">🎨 Formato de salida</div>', unsafe_allow_html=True)
        do_headers = st.checkbox("Encabezados con color", value=True)
        do_autofit = st.checkbox("Ajustar anchos automáticamente", value=True)
        do_freeze  = st.checkbox("Congelar primera fila", value=True)
        
        # NUEVA: Opción para eliminar duplicados
        st.markdown("---")
        st.markdown('<div class="sidebar-title">🔧 Opciones avanzadas</div>', unsafe_allow_html=True)
        remove_duplicates = st.checkbox("Eliminar filas duplicadas", value=False)
        remove_empty_rows = st.checkbox("Eliminar filas vacías", value=False)


# ── Contenido principal ───────────────────────────────────────────────────────
if uploaded is None:
    st.info("⬆  Sube un archivo Excel en el panel izquierdo para comenzar.")
    st.stop()

# Leer datos
try:
    nrows_arg = int(nrows_val) if nrows_val > 0 else None
    df_raw = pd.read_excel(
        uploaded,
        sheet_name=sheet_input,
        header=int(header_row) - 1,
        skiprows=int(skip_rows) if skip_rows > 0 else None,
        nrows=nrows_arg
    )
    
    # Aplicar limpieza si está habilitada
    if remove_empty_rows:
        df_raw = df_raw.dropna(how='all')
    
except Exception as e:
    st.error(f"❌ Error al leer el archivo: {e}")
    st.stop()

all_columns = list(df_raw.columns)
numeric_cols = detect_numeric_columns(df_raw)

# ── Métricas rápidas
c1, c2, c3, c4 = st.columns(4)
c1.metric("Filas", f"{len(df_raw):,}")
c2.metric("Columnas", f"{len(df_raw.columns):,}")
c3.metric("Hoja leída", sheet_input)
c4.metric("Celdas totales", f"{len(df_raw) * len(df_raw.columns):,}")

st.markdown("---")

# ── Tabs principales
tab1, tab2, tab3, tab4 = st.tabs([
    "01  DATOS ORIGINALES", 
    "02  CONFIGURACIÓN", 
    "03  RESULTADO",
    "04  ESTADÍSTICAS"
])

# ── Tab 1: Vista previa
with tab1:
    st.markdown('<div class="panel-label">Vista previa — primeras 50 filas</div>', unsafe_allow_html=True)

    # Filtros avanzados
    with st.expander("🔍 Filtros avanzados (múltiples condiciones)"):
        num_filters = st.number_input("Número de filtros", min_value=0, max_value=5, value=0, step=1)
        
        filters = []
        for i in range(int(num_filters)):
            st.markdown(f"**Filtro {i+1}**")
            fc1, fc2, fc3 = st.columns([2, 1, 2])
            with fc1:
                filter_col = st.selectbox(f"Columna {i+1}", all_columns, key=f"fcol_{i}")
            with fc2:
                filter_op = st.selectbox(
                    f"Condición {i+1}", 
                    ["==", "!=", ">", ">=", "<", "<=", "contains", "not_contains"], 
                    key=f"fop_{i}"
                )
            with fc3:
                filter_val = st.text_input(f"Valor {i+1}", key=f"fval_{i}")
            
            filters.append({"col": filter_col, "op": filter_op, "val": filter_val})

    df_preview = df_raw.copy()
    
    # Aplicar filtros
    if filters:
        df_preview = apply_advanced_filters(df_preview, filters)
        if len(df_preview) < len(df_raw):
            st.caption(f"✅ Mostrando {len(df_preview):,} de {len(df_raw):,} filas tras aplicar filtros.")

    st.dataframe(df_preview.head(50), use_container_width=True, height=320)
    
    # Mostrar columnas numéricas detectadas
    if numeric_cols:
        st.caption(f"📊 Columnas numéricas detectadas: {', '.join(numeric_cols[:10])}")

# ── Tab 2: Configuración
with tab2:
    st.markdown('<div class="panel-label">Modo de transposición</div>', unsafe_allow_html=True)
    mode = st.radio(
        "modo",
        ["Ancho → Largo (melt)", "Largo → Ancho (pivot)", "Transponer todo (.T)"],
        label_visibility="collapsed"
    )

    st.markdown("---")
    params = {}

    if mode == "Ancho → Largo (melt)":
        st.markdown('<div class="panel-label">Columnas identificadoras (no se trasponen)</div>', unsafe_allow_html=True)
        params["id_cols"] = st.multiselect(
            "id_cols", all_columns, label_visibility="collapsed",
            placeholder="Selecciona columnas ID (ej: ID, Nombre, Region...)"
        )
        remaining = [c for c in all_columns if c not in params["id_cols"]]
        st.markdown('<div class="panel-label">Columnas a trasponer</div>', unsafe_allow_html=True)
        params["val_cols"] = st.multiselect(
            "val_cols", remaining, default=remaining, label_visibility="collapsed",
            placeholder="Selecciona columnas a convertir en filas..."
        )
        mc1, mc2 = st.columns(2)
        with mc1:
            params["var_name"] = st.text_input("Nombre columna 'variable'", value="Variable")
        with mc2:
            params["val_name"] = st.text_input("Nombre columna 'valor'", value="Valor")

    elif mode == "Largo → Ancho (pivot)":
        pc1, pc2, pc3 = st.columns(3)
        with pc1:
            params["pivot_index"] = st.selectbox("Columna índice (ID)", all_columns)
        with pc2:
            params["pivot_cols"] = st.selectbox("Columna de variables", all_columns)
        with pc3:
            params["pivot_vals"] = st.selectbox("Columna de valores", all_columns)
        popt1, popt2 = st.columns(2)
        with popt1:
            params["agg_func"] = st.selectbox("Función de agregación",
                                               ["sum", "mean", "count", "max", "min", "first", "median"])
        with popt2:
            params["fill_val"] = st.text_input("Rellenar vacíos con", value="0")

    elif mode == "Transponer todo (.T)":
        st.markdown('<div class="panel-label">Columnas a incluir (vacío = todas)</div>', unsafe_allow_html=True)
        params["val_cols"] = st.multiselect(
            "cols_T", all_columns, label_visibility="collapsed",
            placeholder="Vacío = usa todas las columnas"
        )

    st.markdown("---")
    st.markdown('<div class="panel-label">Ordenamiento del resultado</div>', unsafe_allow_html=True)
    so1, so2 = st.columns([3, 1])
    with so1:
        sort_col = st.text_input("Ordenar por columna (nombre exacto, dejar vacío = no ordenar)", value="")
    with so2:
        sort_asc = st.selectbox("Orden", ["Ascendente", "Descendente"])

    params["sort_col"] = sort_col.strip() if sort_col.strip() else None
    params["sort_asc"] = sort_asc == "Ascendente"
    params["filters"] = filters if 'filters' in locals() else []
    params["remove_duplicates"] = remove_duplicates

    st.markdown("---")
    run_btn = st.button("⇄  Aplicar transposición", use_container_width=True)

# ── Tab 3: Resultado
with tab3:
    if "df_result" not in st.session_state:
        st.info("⚙️ Configura los parámetros en la pestaña 02 y haz clic en 'Aplicar transposición'.")
    else:
        df_res = st.session_state["df_result"]
        st.markdown('<div class="panel-label">Vista previa del resultado</div>', unsafe_allow_html=True)

        rm1, rm2, rm3, rm4 = st.columns(4)
        rm1.metric("Filas resultado", f"{len(df_res):,}")
        rm2.metric("Columnas resultado", f"{len(df_res.columns):,}")
        rm3.metric("Modo aplicado", st.session_state.get("applied_mode", "—"))
        
        # Calcular reducción/expansión
        orig_cells = len(df_raw) * len(df_raw.columns)
        result_cells = len(df_res) * len(df_res.columns)
        change_pct = ((result_cells - orig_cells) / orig_cells * 100) if orig_cells > 0 else 0
        rm4.metric("Cambio en datos", f"{change_pct:+.1f}%")

        st.dataframe(df_res.head(100), use_container_width=True, height=380)

        excel_bytes = to_excel_bytes(
            df_res, out_sheet, include_idx, do_headers, do_autofit, do_freeze
        )
        fname = out_filename if out_filename.endswith(".xlsx") else out_filename + ".xlsx"
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.download_button(
                label="↓  Descargar resultado Excel",
                data=excel_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col2:
            # Opción para exportar CSV
            csv_data = df_res.to_csv(index=include_idx).encode('utf-8')
            st.download_button(
                label="↓  CSV",
                data=csv_data,
                file_name=fname.replace(".xlsx", ".csv"),
                mime="text/csv",
                use_container_width=True
            )

# ── Tab 4: Estadísticas (NUEVO)
with tab4:
    if "df_result" not in st.session_state:
        st.info("⚙️ Procesa datos en la pestaña 02 primero.")
    else:
        df_res = st.session_state["df_result"]
        st.markdown('<div class="panel-label">📊 Análisis estadístico del resultado</div>', unsafe_allow_html=True)
        
        # Estadísticas generales
        st.markdown("**Resumen general**")
        stat1, stat2, stat3, stat4 = st.columns(4)
        stat1.metric("Total de celdas", f"{len(df_res) * len(df_res.columns):,}")
        stat2.metric("Valores únicos (promedio)", f"{df_res.nunique().mean():.1f}")
        
        # Calcular valores nulos
        total_nulls = df_res.isnull().sum().sum()
        null_pct = (total_nulls / (len(df_res) * len(df_res.columns)) * 100) if len(df_res) > 0 else 0
        stat3.metric("Valores nulos", f"{total_nulls:,} ({null_pct:.1f}%)")
        stat4.metric("Filas completas", f"{df_res.dropna().shape[0]:,}")
        
        st.markdown("---")
        
        # Análisis por columna
        st.markdown("**Análisis por columna**")
        numeric_result_cols = detect_numeric_columns(df_res)
        
        if numeric_result_cols:
            selected_col = st.selectbox("Selecciona columna numérica para análisis", numeric_result_cols)
            
            desc_stats = df_res[selected_col].describe()
            
            dcol1, dcol2, dcol3, dcol4, dcol5 = st.columns(5)
            dcol1.metric("Media", f"{desc_stats['mean']:.2f}")
            dcol2.metric("Mediana", f"{df_res[selected_col].median():.2f}")
            dcol3.metric("Desv. Std", f"{desc_stats['std']:.2f}")
            dcol4.metric("Mínimo", f"{desc_stats['min']:.2f}")
            dcol5.metric("Máximo", f"{desc_stats['max']:.2f}")
        else:
            st.info("No hay columnas numéricas en el resultado para análisis estadístico.")
        
        # Información de tipos de datos
        st.markdown("---")
        st.markdown("**Tipos de datos**")
        dtype_info = df_res.dtypes.value_counts()
        for dtype, count in dtype_info.items():
            st.caption(f"• **{dtype}**: {count} columna(s)")

# ── Ejecutar transposición
if run_btn:
    try:
        df_work = df_raw.copy()

        # Aplicar filtros si existen
        if params.get("filters"):
            df_work = apply_advanced_filters(df_work, params["filters"])

        df_result = process_dataframe(df_work, mode, params)
        
        # Eliminar duplicados si está habilitado
        if params.get("remove_duplicates"):
            before_dup = len(df_result)
            df_result = df_result.drop_duplicates()
            after_dup = len(df_result)
            if before_dup > after_dup:
                st.info(f"🔧 Eliminadas {before_dup - after_dup:,} filas duplicadas")

        # Ordenar
        sc = params.get("sort_col")
        if sc and sc in df_result.columns:
            df_result = df_result.sort_values(sc, ascending=params.get("sort_asc", True))

        st.session_state["df_result"]    = df_result
        st.session_state["applied_mode"] = mode.split("(")[0].strip()
        st.success(f"✅ Transposición completada — {len(df_result):,} filas × {len(df_result.columns):,} columnas")
        st.rerun()

    except Exception as e:
        st.error(f"❌ Error al procesar: {e}")
